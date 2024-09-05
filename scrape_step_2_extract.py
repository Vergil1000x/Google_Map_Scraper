import asyncio
import openpyxl
from bs4 import BeautifulSoup
from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
)


async def fetch_page_content(semaphore, browser, url, worksheet_write):
    async with semaphore:
        page = await browser.new_page()
        try:
            # Navigate to the URL and wait for network idle to ensure full load
            await page.goto(url, wait_until="networkidle")
            await asyncio.sleep(5)  # Optional: Adjust based on loading speed

            # Get page content and parse with BeautifulSoup
            content = await page.content()
            soup = BeautifulSoup(content, "html.parser")

            # Initialize a list to hold extracted data
            string_list = []

            # Define selectors for elements to extract
            selectors = [
                ("DUwDvf lfPIob", "class"),
                ("F7nice", "class"),
                ("Copy address", "data-tooltip"),
                ("Open booking link", "data-tooltip"),
                ("Open website", "data-tooltip"),
                ("Copy phone number", "data-tooltip"),
            ]

            # Extract data based on selectors
            for selector, selector_type in selectors:
                element = (
                    soup.find(attrs={selector_type: selector})
                    if selector_type != "class"
                    else soup.find(class_=selector)
                )
                if element:
                    if selector_type == "data-tooltip" and element.has_attr("href"):
                        href = element["href"]
                        string_list.append(href)
                        print(f"Extracted href: {href}")
                    else:
                        text = element.get_text(strip=True)
                        string_list.append(text)
                        print(f"Extracted text: {text}")
                else:
                    string_list.append("Not Available")
                    print(f"Element with {selector_type} = '{selector}' not found.")

            # Write data to Excel (synchronously to avoid conflicts)
            worksheet_write.append(string_list)

        except PlaywrightTimeoutError:
            print(f"Timeout error while loading {url}")
            worksheet_write.append(["Timeout error"])
        except Exception as e:
            print(f"Error loading {url}: {e}")
            worksheet_write.append([f"Error: {e}"])
        finally:
            await page.close()


async def main(
    input_file_path=r"fx.xlsx",
    output_file_path=r"fxout.xlsx",
):
    # Load the workbook and worksheet for reading URLs
    workbook_read = openpyxl.load_workbook(input_file_path)
    worksheet_read = workbook_read.active

    # Extract website URLs from the Excel file
    urls = [
        row[0]
        for row in worksheet_read.iter_rows(min_row=1, max_col=1, values_only=True)
        if row[0] and isinstance(row[0], str) and row[0].startswith("http")
    ]

    workbook_read.close()

    # Load the workbook and worksheet for saving data
    workbook_write = openpyxl.load_workbook(output_file_path)
    worksheet_write = workbook_write.active

    # Use `async_playwright` context manager to manage Playwright resources
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        # ************************************* Control concurrency level *************************************
        semaphore = asyncio.Semaphore(3)
        tasks = [
            fetch_page_content(semaphore, browser, url, worksheet_write) for url in urls
        ]
        await asyncio.gather(*tasks)
        await browser.close()

        # Save the workbook after all tasks are complete
        try:
            workbook_write.save(output_file_path)
            print("Data saved successfully to Excel.")
        except Exception as e:
            print(f"Error saving workbook: {e}")

    workbook_write.close()


if __name__ == "__main__":
    asyncio.run(main())
