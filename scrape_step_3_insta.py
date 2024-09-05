import os
import asyncio
import openpyxl
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup


async def main(
    output_file_path=r"fxout.xlsx",
    insta_file_path=r"fxinsta2.xlsx",
):
    # Load the workbook and worksheet for reading URLs
    workbook_read = openpyxl.load_workbook(output_file_path, read_only=True)
    worksheet_read = workbook_read.active

    # Extract website URLs from the Excel file
    website_urls = [
        row[i]
        for row in worksheet_read.iter_rows(
            min_row=1, min_col=4, max_col=5, values_only=True
        )
        for i in range(len(row))
        if row[i] and isinstance(row[i], str) and row[i].startswith("http")
    ]

    workbook_read.close()

    # Ensure the output workbook exists or create a new one
    if not os.path.exists(insta_file_path):
        workbook_write = openpyxl.Workbook()
        worksheet_write = workbook_write.active
        worksheet_write.title = "Instagram Links"
    else:
        workbook_write = openpyxl.load_workbook(insta_file_path)
        worksheet_write = workbook_write.active

    async def fetch_instagram_links_with_playwright(url, semaphore, browser):
        async with semaphore:
            page = await browser.new_page()
            try:
                await page.goto(url, timeout=30000)  # 30 seconds timeout
                content = await page.content()
                soup = BeautifulSoup(content, "html.parser")
                instagram_links = [
                    a["href"]
                    for a in soup.find_all("a", href=True)
                    if "instagram.com" in a["href"]
                ]
                if instagram_links:
                    print(f"Instagram links found on {url}:")
                    for link in instagram_links:
                        print(link)
                        worksheet_write.append([link])
            except Exception as e:
                print(f"Error fetching data from {url}: {str(e)}")
            finally:
                await page.close()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        # ************************************* Control the number of concurrent tasks *************************************
        semaphore = asyncio.Semaphore(5)
        tasks = [
            fetch_instagram_links_with_playwright(url, semaphore, browser)
            for url in website_urls
        ]
        await asyncio.gather(*tasks)
        await browser.close()

    # Save the workbook after all tasks are complete
    try:
        workbook_write.save(insta_file_path)
        print(f"Data saved successfully to {insta_file_path}.")
    except Exception as e:
        print(f"Error saving workbook: {e}")
    finally:
        workbook_write.close()


# Run the main function
if __name__ == "__main__":
    asyncio.run(main())
